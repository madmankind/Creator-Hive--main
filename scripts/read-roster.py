#!/usr/bin/env python3
import json, sys, os
try:
    import openpyxl
except ImportError:
    print('{"error":"pip3 install openpyxl"}', file=sys.stderr)
    sys.exit(1)
path = "/Users/ajil/Desktop/Creator Hive/Talent List MAIN/CH_MasterRoster_v1_MAR17.xlsx"
if not os.path.exists(path):
    print(json.dumps({"error":"File not found"}), file=sys.stderr)
    sys.exit(1)
wb = openpyxl.load_workbook(path, data_only=True)
out = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    h = [str(x or "").strip() for x in rows[0]] if rows else []
    out[sn] = {"headers": h, "rows": [dict(zip(h, [str(v) if v is not None else "" for v in r])) for r in rows[1:]]}
wb.close()
print(json.dumps(out, indent=2, default=str))
